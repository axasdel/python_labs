import argparse
import json
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

import sys
sys.path.append(r"C:\Users\user\Desktop\python_labs\src")


def csv2xlsx(csv_path_1, xlsx_path_2):
    csv_newpath = Path(csv_path_1)
    xlsx_newpath = Path(xlsx_path_2)

    if not csv_newpath.exists():
        raise FileNotFoundError("Файла csv не существует")

    workbook = Workbook()  # создание книги эксель
    work_sheet = workbook.active  # активный лист
    work_sheet.title = "Sheet1"

    column_width = {}  # для макс ширины столбца

    with csv_newpath.open("r", encoding="utf-8") as cf:
        csv_reading = csv.reader(cf)
        for row in csv_reading:
            work_sheet.append(row)
            for column_ind, value in enumerate(row, start=1):  # enumerate для получения (индекс, значение)
                cell_length = len(value)  # длина значения
                if cell_length > column_width.get(column_ind, 0):  # случай, если длина значения > текущей макс ширины
                    column_width[column_ind] = cell_length

    for column_ind, width in column_width.items():  # для индекса, ширины в словаре
        letter = get_column_letter(column_ind)  # исходя из индекса, присваеваем столбцам экселевские буквы
        work_sheet.column_dimensions[letter].width = ((width + 2) if (width + 2) >= 8 else 8)  # по 1 отступу с обеих сторон

    workbook.save(xlsx_newpath)


def json2csv(json_path, csv_path):
    json_newpath = Path(json_path)
    csv_newpath = Path(csv_path)

    if not json_newpath.exists():
        raise FileNotFoundError("Файл отсутствует")

    with json_newpath.open("r", encoding="utf-8") as jf:
        json_text = json.load(jf)

    if json_text is None:
        raise ValueError("Файл пуст")
    if len(json_text) == 0:
        raise ValueError("Файл пуст")
    if type(json_text) != list:
        raise ValueError("JSON не список")
    if not all(isinstance(element, dict) for element in json_text):  # все ли элементы - словари
        raise ValueError("JSON - список словарей")

    fieldnames = list(json_text[0].keys())  # преобразование ключей 1ого словаря в заголовок для будующего csv
    with csv_newpath.open("w", encoding="utf-8", newline="") as cf:
        csv_writer = csv.DictWriter(cf, fieldnames=fieldnames, extrasaction="ignore")
        csv_writer.writeheader()  # игнорирование лишних ключей
        csv_writer.writerows(json_text)


def csv2json(csv_path, json_path):
    csv_newpath = Path(csv_path)
    json_newpath = Path(json_path)

    if not csv_newpath.exists():
        raise FileNotFoundError("Файл отсутствует")

    with csv_newpath.open("r", encoding="utf-8", newline="") as cf:
        csv_text = csv.DictReader(cf, delimiter=",")

        if not csv_text.fieldnames:
            raise ValueError("Отсутствуют заголовки")

        text = [dict(row) for row in csv_text]
        if len(text) == 0:
            raise ValueError("Файл пуст")

    with json_newpath.open("w", encoding="utf-8") as jf:
        json.dump(text, jf, ensure_ascii=False, indent=2)
        # отступы


def main():
    parser = argparse.ArgumentParser(description='CLI для конвертации файла')
    subparsers = parser.add_subparsers(dest='command', help='Доступные команды')


    #подкоманда json2csv
    json2csv_parser = subparsers.add_parser('json2csv', help='JSON -> CSV')
    json2csv_parser.add_argument('--in', required=True, dest='input_file', help='Входной JSON')
    json2csv_parser.add_argument('--out', required=True, dest='output_file', help='Выходной CSV')


    #подкоманда csv2json
    csv2json_parser = subparsers.add_parser('csv2json', help='CSV -> JSON')
    csv2json_parser.add_argument('--in', required=True, dest='input_file', help='Вхдной CSV')
    csv2json_parser.add_argument('--out', required=True, dest='output_file', help='Выходной JSON')


    #подкоманда csv2xlsx
    csv2xlsx_parser = subparsers.add_parser('csv2xlsx', help='CSV -> XLSX')
    csv2xlsx_parser.add_argument('--in', required=True, dest='input_file', help='Входной CSV')
    csv2xlsx_parser.add_argument('--out', required=True, dest='output_file', help='Выходной XLSX')


    args = parser.parse_args()


    try:
        if args.command == 'json2csv':
            json2csv(args.input_file, args.output_file)
        elif args.command == 'csv2json':
            csv2json(args.input_file, args.output_file)
        elif args.command == 'csv2xlsx':
            csv2xlsx(args.input_file, args.output_file)
    except Exception as e:
        parser.error


if __name__ == '__main__':
    main()