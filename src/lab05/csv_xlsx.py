import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from pathlib import Path


def csv_to_xlsx(csv_path_1, xlsx_path_2):
    csv_newpath = Path(csv_path_1)
    xlsx_newpath = Path(xlsx_path_2)

    if not csv_newpath.exists():
        raise FileNotFoundError('Файла csv не существует')
    
    workbook = Workbook()
    work_sheet = workbook.active
    work_sheet.title = 'Sheet1'

    collumn_width = {}
    for collumn_ind, width in collumn_width.items():
        letter = get_column_letter(collumn_ind)
        work_sheet.column_dimensions[letter].width = (width + 2) if \
        (width + 2) >= 8 else 8

    with csv_newpath.open('r', encoding='utf-8') as cf:
        csv_reading = csv.reader(cf)
        for row in csv_reading:
            work_sheet.append(row)
            for collumn_ind, value in enumerate(row, start = 1):
                cell_length = len(value)
                if cell_length > collumn_width.get(collumn_ind, 0):
                    collumn_width[collumn_ind] = cell_length

    workbook.save(xlsx_newpath)


csv_path = 'lab05/data/samples/people.csv'
xlsx_path = 'lab05/data/result/res_people.csv'
csv_to_xlsx(csv_path, xlsx_path)

